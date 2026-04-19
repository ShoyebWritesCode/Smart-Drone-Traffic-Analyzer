import pandas as pd
import os
from datetime import datetime

def generate_report(counts, total_frames, fps, output_dir, events=None, processing_duration=0):
    """
    Generate professional CSV and Excel reports including total counts, 
    breakdowns, processing duration, and detailed detection logs.
    """
    os.makedirs(output_dir, exist_ok=True)
    
    # Vehicle class mapping
    class_names = {
        2: "Car",
        3: "Motorcycle",
        5: "Bus",
        7: "Truck"
    }
    
    total_vehicles = sum(counts.values())
    video_duration = round(total_frames / fps, 2) if fps > 0 else 0
    
    # 1. SUMMARY DATA (Counts & Analytics)
    summary_data = []
    for cls_id, count in counts.items():
        name = class_names.get(int(cls_id), f"Class {cls_id}")
        summary_data.append({
            "Vehicle Type": name,
            "Count": count,
            "Percentage": f"{(count/total_vehicles*100):.1f}%" if total_vehicles > 0 else "0%"
        })
    df_summary = pd.DataFrame(summary_data)
    
    # 2. METADATA (Performance & Stats)
    metadata_data = [{
        "Report Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Total Vehicles Detected": total_vehicles,
        "Processing Duration (s)": processing_duration,
        "Video Duration (s)": video_duration,
        "Total Frames": total_frames,
        "FPS": round(fps, 2)
    }]
    df_metadata = pd.DataFrame(metadata_data)
    
    # 3. DETAILED DETECTION LOG (Events)
    if events:
        log_data = []
        for event in events:
            total_seconds = float(event['timestamp'])
            mins = int(total_seconds // 60)
            secs = int(total_seconds % 60)
            millis = int(round((total_seconds % 1) * 1000))
            formatted_time = f"{mins:02d}:{secs:02d}.{millis:03d}"
            
            log_data.append({
                "Timestamp (MM:SS.ms)": formatted_time,
                "Frame Number": event['frame'],
                "Vehicle Type": class_names.get(int(event['class_id']), f"Class {event['class_id']}"),
                "Unique Track ID": event['track_id'],
                "Confidence": f"{event.get('confidence', 0)}%"
            })
        df_log = pd.DataFrame(log_data)
    else:
        df_log = pd.DataFrame(columns=["Timestamp (MM:SS.ms)", "Frame Number", "Vehicle Type", "Unique Track ID", "Confidence"])

    # Output paths
    csv_path = os.path.join(output_dir, "report.csv")
    excel_path = os.path.join(output_dir, "report.xlsx")
    
    # SAVE CSV: We'll save the detailed log as the primary CSV
    # but include the summary as a header-like block
    with open(csv_path, 'w') as f:
        f.write("# SMART DRONE TRAFFIC ANALYZER - SUMMARY REPORT\n")
        f.write(f"# Generated: {metadata_data[0]['Report Generated']}\n")
        f.write(f"# Total Vehicles: {total_vehicles}\n")
        
        # Add individual class counts to CSV header
        for cls_id, count in counts.items():
            name = class_names.get(int(cls_id), f"Class {cls_id}")
            f.write(f"# -> {name}: {count}\n")
            
        f.write(f"# AI Processing Duration: {processing_duration}s\n\n")
        df_log.to_csv(f, index=False)
    
    # SAVE EXCEL: Separate sheets for a professional look
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name="Vehicle Summary", index=False)
        df_metadata.to_excel(writer, sheet_name="Performance Stats", index=False)
        df_log.to_excel(writer, sheet_name="Detection Log", index=False)
        
    # Apply beautiful formatting to the Excel file using openpyxl
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        
        wb = load_workbook(excel_path)
        
        # Cyan neon header fill
        header_fill = PatternFill(start_color="22D3EE", end_color="22D3EE", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 4)
                ws.column_dimensions[column].width = adjusted_width

        wb.save(excel_path)
    except Exception as e:
        print(f"Warning: Could not format Excel file: {e}")
        
    print(f"Detailed reports (CSV & Excel) generated in {output_dir}")
